#!/usr/bin/env python3
"""Generate 公安经侦 objective question bank (20 questions) into 题库导入模板.xlsx."""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

HEADERS = ["题干", "题型", "A", "B", "C", "D", "E", "F", "答案", "解析", "知识点", "难度", "分值", "警种"]

QUESTIONS = [
    {
        "stem": "公安机关经济犯罪侦查部门的简称是？",
        "type": "单选",
        "options": ["刑侦", "经侦", "治安", "网安"],
        "answer": "B",
        "explanation": "经济犯罪侦查部门简称“经侦”。",
        "knowledge": "经侦概述",
        "difficulty": 1,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "经侦部门主要打击危害社会主义市场经济秩序的哪类犯罪？",
        "type": "单选",
        "options": ["经济犯罪", "暴力犯罪", "职务犯罪", "环境犯罪"],
        "answer": "A",
        "explanation": "经侦部门以打击经济犯罪为主要职责。",
        "knowledge": "经侦职责",
        "difficulty": 1,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "合同诈骗罪规定在《刑法》第几条？",
        "type": "单选",
        "options": ["192", "205", "224", "266"],
        "answer": "C",
        "explanation": "合同诈骗罪规定在刑法第224条。",
        "knowledge": "合同诈骗罪",
        "difficulty": 2,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "非法吸收公众存款罪侵犯的客体是国家的什么？",
        "type": "单选",
        "options": ["公共安全", "金融管理秩序", "社会管理秩序", "公民人身权利"],
        "answer": "B",
        "explanation": "该罪侵犯客体为国家金融管理秩序。",
        "knowledge": "非法吸收公众存款罪",
        "difficulty": 2,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "组织、领导传销活动罪规定在《刑法》第几条？",
        "type": "单选",
        "options": ["140", "224条之一", "266", "312"],
        "answer": "B",
        "explanation": "组织、领导传销活动罪规定在刑法第224条之一。",
        "knowledge": "传销犯罪",
        "difficulty": 3,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "《反洗钱法》规定，金融机构应当履行什么义务？",
        "type": "单选",
        "options": ["反洗钱", "反恐怖融资以外的全部监管", "公开客户信息", "代客户保管密码"],
        "answer": "A",
        "explanation": "金融机构应建立客户身份识别等制度，履行反洗钱义务。",
        "knowledge": "反洗钱",
        "difficulty": 2,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "对犯罪嫌疑人存款等财产进行冻结，冻结期限最长为？",
        "type": "单选",
        "options": ["3个月", "6个月", "12个月", "24个月"],
        "answer": "B",
        "explanation": "冻结存款等财产的期限最长为六个月。",
        "knowledge": "侦查措施",
        "difficulty": 2,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "虚开增值税专用发票罪规定在《刑法》第几条？",
        "type": "单选",
        "options": ["140", "205", "224", "266"],
        "answer": "B",
        "explanation": "虚开增值税专用发票罪规定在刑法第205条。",
        "knowledge": "涉税犯罪",
        "difficulty": 2,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "职务侵占罪的主体通常为？",
        "type": "单选",
        "options": ["国家工作人员", "非国家工作人员", "军人", "任意公民"],
        "answer": "B",
        "explanation": "职务侵占罪主体为公司、企业或其他单位中从事公务以外的人员。",
        "knowledge": "职务侵占罪",
        "difficulty": 2,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "非法吸收公众存款罪，个人犯罪数额达到多少以上应予立案追诉？",
        "type": "单选",
        "options": ["10万元", "20万元", "50万元", "100万元"],
        "answer": "B",
        "explanation": "个人非法吸收或变相吸收公众存款数额在20万元以上应予立案追诉。",
        "knowledge": "立案标准",
        "difficulty": 3,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "生产、销售伪劣产品罪规定在《刑法》第几条？",
        "type": "单选",
        "options": ["140", "205", "224", "266"],
        "answer": "A",
        "explanation": "生产、销售伪劣产品罪规定在刑法第140条。",
        "knowledge": "伪劣产品犯罪",
        "difficulty": 2,
        "points": 2,
        "category": "经侦",
    },
    {
        "stem": "下列哪些属于经侦部门常见打击的犯罪类型？",
        "type": "多选",
        "options": ["合同诈骗", "非法吸收公众存款", "交通肇事", "职务侵占"],
        "answer": "A、B、D",
        "explanation": "交通肇事属于交通或刑侦管辖，其余均为典型经济犯罪。",
        "knowledge": "经侦管辖",
        "difficulty": 2,
        "points": 3,
        "category": "经侦",
    },
    {
        "stem": "洗钱罪的上游犯罪包括哪些？",
        "type": "多选",
        "options": ["毒品犯罪", "走私犯罪", "妨害公务罪", "贪污贿赂犯罪"],
        "answer": "A、B、D",
        "explanation": "洗钱罪上游犯罪包括毒品、走私、贪污贿赂等七类犯罪，不含妨害公务罪。",
        "knowledge": "洗钱罪",
        "difficulty": 3,
        "points": 3,
        "category": "经侦",
    },
    {
        "stem": "金融机构反洗钱制度通常包括哪些内容？",
        "type": "多选",
        "options": ["客户身份识别", "客户身份资料保存", "交易记录保存", "公开客户密码"],
        "answer": "A、B、C",
        "explanation": "反洗钱法要求建立客户身份识别和资料、交易记录保存制度。",
        "knowledge": "反洗钱",
        "difficulty": 2,
        "points": 3,
        "category": "经侦",
    },
    {
        "stem": "公安机关办理经济犯罪案件，证明标准应坚持哪些要求？",
        "type": "多选",
        "options": ["事实清楚", "证据确实、充分", "程序合法", "无需审查证据"],
        "answer": "A、B、C",
        "explanation": "办理刑事案件应事实清楚、证据确实充分，并依法保障程序合法。",
        "knowledge": "证明标准",
        "difficulty": 2,
        "points": 3,
        "category": "经侦",
    },
    {
        "stem": "经侦民警在办理涉企案件时，应当落实哪些司法政策？",
        "type": "多选",
        "options": ["少捕", "慎诉", "慎押", "一律从重处罚"],
        "answer": "A、B、C",
        "explanation": "涉企案件应落实少捕慎诉慎押等政策，保障市场主体合法权益。",
        "knowledge": "涉企司法政策",
        "difficulty": 2,
        "points": 3,
        "category": "经侦",
    },
    {
        "stem": "经侦部门在侦查非法集资案件时，应当准确区分非法集资与哪些行为？",
        "type": "多选",
        "options": ["民间借贷", "合法融资", "盗窃", "诈骗"],
        "answer": "A、B",
        "explanation": "非法集资案件需与民间借贷、合法融资等合法行为相区分。",
        "knowledge": "非法集资",
        "difficulty": 3,
        "points": 3,
        "category": "经侦",
    },
    {
        "stem": "办理跨境经济犯罪案件，公安机关可以依法开展哪些工作？",
        "type": "多选",
        "options": ["国际执法合作", "国际警务合作", "追查违法所得", "直接境外抓人无需手续"],
        "answer": "A、B、C",
        "explanation": "可依法开展国际执法合作并追查违法所得，境外执法须依法进行。",
        "knowledge": "跨境案件",
        "difficulty": 3,
        "points": 3,
        "category": "经侦",
    },
    {
        "stem": "判断题：公安机关经济犯罪侦查部门简称“经侦”。",
        "type": "判断",
        "options": ["正确", "错误"],
        "answer": "A",
        "explanation": "经济犯罪侦查部门简称经侦，表述正确。",
        "knowledge": "经侦概述",
        "difficulty": 1,
        "points": 1,
        "category": "经侦",
    },
    {
        "stem": "判断题：查询、冻结犯罪嫌疑人存款、汇款等财产，可以不经县级以上公安机关负责人批准。",
        "type": "判断",
        "options": ["正确", "错误"],
        "answer": "B",
        "explanation": "查询、冻结上述财产应当经县级以上公安机关负责人批准。",
        "knowledge": "侦查措施",
        "difficulty": 2,
        "points": 1,
        "category": "经侦",
    },
]


def row_from_question(q: dict) -> list:
    opts = q["options"]
    padded = opts + [None] * (6 - len(opts))
    return [
        q["stem"],
        q["type"],
        padded[0],
        padded[1],
        padded[2],
        padded[3],
        padded[4],
        padded[5],
        q["answer"],
        q["explanation"],
        q["knowledge"],
        q["difficulty"],
        q["points"],
        q.get("category"),
    ]


def main() -> int:
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Desktop" / "题库导入模板.xlsx"
    template = Path(__file__).resolve().parents[1] / "templates" / "题库导入模板.xlsx"
    if template.is_file():
        wb = openpyxl.load_workbook(template)
        ws = wb["题库导入"]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "题库导入"
        ws.append(HEADERS)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for q in QUESTIONS:
        ws.append(row_from_question(q))

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    types = {}
    for q in QUESTIONS:
        types[q["type"]] = types.get(q["type"], 0) + 1
    print(f"Wrote {out}")
    print(f"Questions: {len(QUESTIONS)} — " + ", ".join(f"{k} {v}" for k, v in sorted(types.items())))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
